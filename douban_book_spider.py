import time
import random
import requests
from urllib.parse import quote
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor


hds = [
    {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
    {'User-Agent': ('Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) '
     'Chrome/17.0.963.12 Safari/535.11')},
    {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}
]

sess = requests.session()
def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0
    retry_limit = 50
    # while True:
    while page_num < 10:  # 只抓取前10页
        url = f'https://www.douban.com/tag/{quote(book_tag)}/book?start={page_num * 15}'
        time.sleep(random.random() * 5)

        try:
            response = sess.get(url, headers=hds[page_num % len(hds)])
            source_code = response.content
            plain_text = str(source_code, 'utf-8')
        except Exception as e:
            print(e)
            continue

        soup = BeautifulSoup(plain_text, 'html.parser')
        list_soup = soup.find('div', {'class': 'mod book-list'})

        try_times += 1
        if list_soup is None and try_times < retry_limit:
            continue
        elif list_soup is None or len(list_soup) <= 1:
            break

        for book_info in list_soup.find_all('dd'):
            title = book_info.find('a', {'class': 'title'}).text.strip()
            desc = book_info.find('div', {'class': 'desc'}).text.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class': 'title'})['href']

            author_info = '作者/译者： ' + '/'.join(desc_list[:-3]) if len(desc_list) > 3 else '作者/译者： 暂无'
            if len(desc_list) > 3:
                print(desc_list)
            else:
                print('作者/译者： 暂无')

            pub_info = '出版信息： ' + '/'.join(desc_list[-3:]) if len(desc_list) >= 3 else '出版信息： 暂无'
            if len(desc_list) >= 3:
                print(desc_list)
            else:
                print('出版信息： 暂无')

            rating_span = book_info.find('span', {'class': 'rating_nums'})
            rating = rating_span.text.strip() if rating_span else '0.0'
            if rating_span:
                print(rating_span)
            else:
                print('0.0')
            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except Exception as e:
                people_num = '0'

            book_list.append([title, rating, people_num, author_info, pub_info])
            print(f"Book: {title}, Rating: {rating}, People Num: {people_num}, Author: {author_info}, Pub Info: {pub_info}")
            try_times = 0
        page_num += 1
        print(f'Downloading Information From Page {page_num}')
    return book_list


def get_people_num(url):
    try:
        response = sess.get(url, headers=hds[random.randint(0, len(hds) - 1)])
        source_code = response.content
        plain_text = str(source_code, 'utf-8')
    except Exception as e:
        print(e)
        return '0'

    soup = BeautifulSoup(plain_text, 'html.parser')
    people_num = soup.find('div', {'class': 'rating_sum'}).find_all('span')[1].text.strip()
    return people_num


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        print(f"Starting to spider tag: {book_tag}")
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
        print(f"Finished spidering tag: {book_tag}, found {len(book_list)} books")
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook()
    wb.remove(wb.active)

    for book_tag in book_tag_lists:
        ws = wb.create_sheet(title=book_tag)
        ws.append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        count = 1
        for bl in book_lists[book_tag_lists.index(book_tag)]:
            ws.append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1

        center_alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 55
        ws.column_dimensions['F'].width = 60

        for row in ws.iter_rows(min_row=1):
            row_num = row[0].row
            ws.row_dimensions[row_num].height = 20
            for cell in row:
                cell.alignment = center_alignment

    save_path = 'book_list'
    for book_tag in book_tag_lists:
        save_path += f'-{book_tag}'
    save_path += '.xlsx'

    wb.save(save_path)
    print(f"数据已成功写入 {save_path} 文件。")


if __name__ == '__main__':
    '''
    book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    book_tag_lists = ['数学']
    book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    book_tag_lists = ['商业','理财','管理']  
    book_tag_lists = ['名著']
    book_tag_lists = ['科普','经典','生活','心灵','文学']
    book_tag_lists = ['科幻','思维','金融']
    '''
    book_tag_lists = ['计算机', '编程']
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists, book_tag_lists)