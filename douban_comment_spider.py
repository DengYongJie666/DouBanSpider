import bs4
import requests
from bs4 import BeautifulSoup
import time
import random
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor


sess = requests.session()


def download_page(base_url):
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
            '(HTML, like Gecko) Chrome/129.0.0.0 Safari/537.36'
        )
    }

    page_obj = sess.get(base_url, headers=headers)
    bs4_obj = BeautifulSoup(page_obj.text, 'html.parser')

    if page_obj.status_code == 200:
        total_pages_element = bs4_obj.find("span", {"class": "thispage"})
        if total_pages_element and "data-total-page" in total_pages_element.attrs:
            total_pages = int(total_pages_element["data-total-page"])
            print(f"一共{total_pages}页评论")
        else:
            print("未找到总页数元素或属性")
            total_pages = 1
    else:
        print(f"请求失败，状态码: {page_obj.status_code}")
        total_pages = 1

    bs4_page_obj_list = [bs4_obj]

    for page in range(1, total_pages):
        time.sleep(random.random() * 5)
        start = page * 100
        url = f"{base_url}?start={start}"
        print(f"下载分页 {url}")
        response = sess.get(url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        bs4_page_obj_list.append(soup)

    print(f"共下载了 {len(bs4_page_obj_list)} 页")
    return bs4_page_obj_list


def save_markdown_pubtime_excel(search_keywords, bs4_page_obj_list):
    matching_paragraphs = []
    for bs4_page_obj in bs4_page_obj_list:
        comment_eles = bs4_page_obj.find_all("div", {"class": "reply-doc content"})
        for ele in comment_eles:
            comment_ele = ele.find("div", attrs={"class": "markdown"})
            if comment_ele:
                markdown_text = comment_ele.get_text(strip=True)
                if any(keyword in markdown_text for keyword in search_keywords):
                    pubtime_date = ele.find("span", {"class": "pubtime"})
                    if pubtime_date:
                        pubtime_text = pubtime_date.get_text(strip=True)
                        matching_paragraphs.append([markdown_text, pubtime_text])
    if matching_paragraphs:
        df = pd.DataFrame(matching_paragraphs, columns=['评论', '时间'])
        df.to_excel('excel_test.xlsx', index=False)
        book = load_workbook("excel_test.xlsx")
        sheet = book["Sheet1"]
        center_alignment = Alignment(horizontal="center",vertical="center")

        for col in sheet.columns:
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = 30

        for row in sheet.iter_rows(min_row=1):
            row_num = row[0].row
            sheet.row_dimensions[row_num].height = 20
            for cell in row:
                cell.alignment = center_alignment

        for row in sheet.iter_rows(values_only=True):
            print(row)

        book.save("excel_test.xlsx")
        book.close()
        print("数据已成功写入 excel_test.xlsx 文件。")
    else:
        print("没有找到包含关键字的段落。")


if __name__ == '__main__':
    base_url = "https://www.douban.com/group/topic/292594799/"
    search_keywords = ["申请", "报名"]
    bs4_page_obj_list = download_page(base_url)
    save_markdown_pubtime_excel(search_keywords, bs4_page_obj_list)































